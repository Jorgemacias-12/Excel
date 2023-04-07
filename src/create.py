import pandas as pd
from openpyxl import Workbook, load_workbook

def createFile():
    
    fileName = input('Nombre de archivo: ')

    workbook = Workbook()

    workbook.create_sheet("asistencias")
  
    workbook.active = workbook["asistencias"]

    workbook.save(fileName + ".xlsx")


def insertStudent():
    
    # Lista de asistencia
    # Nombre 
    # Campo de asistencia

    file_name = 'students.xlsx'

    studentName = input("Nombre: ")
    attendance = input("¿Asistio? ")
    
    data = {}

    # studentData = pd.DataFrame({'Nombre: ': studentName, 'Asistencia': attendance})

    writter = pd.ExcelWriter(file_name, engine='openpyxl', mode='a')

    writter.book = load_workbook(file_name)

    studentData.to_excel(writter, 'asistencias', startrow=0, index=False)
    
    writter.save()
